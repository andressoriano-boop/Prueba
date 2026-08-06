from datetime import datetime, date

import requests
import zipfile
import io

import pandas as pd

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

#Se obtiene la fecha de la publicación del ENSU mediante web scraping de una api de la página de INEGI
url_inegi = "https://www.inegi.org.mx/app/api/saladeprensa/api/saladeprensa/ObtenerFechasTabla/v3"

#Se definen los parámetros de la solicitud
payload = {
    "fechaDesde": "0",
    "fechaHasta": "0",
    "titulo": "",
    "idPrograma": 0,
    "ordenarPor": "fecha",
    "ordenarAsc": 0,
    "desde": 0,
    "tomar": 1000,
    "ingles": 0,
    "ambito": -1,
    "tipoNoticia": "1,2,3,4,5,6,7,8"
}

#Se definen los encabezados de la solicitud
headers = {
    "X-Requested-With": "XMLHttpRequest",
    "Referer": "https://www.inegi.org.mx/app/saladeprensa/calendario"
}

#Se realiza la solicitud POST a la API de INEGI
r = requests.post(url_inegi, data=payload, headers=headers)

datos = r.json() #Se obtiene la respuesta en formato JSON

#Se recorre la lista de publicaciones obtenidas para encontrar la fecha de publicación del ENSU
for publicacion in datos:
    fecha = datetime.strptime(publicacion["fecha"], "%d/%m/%Y").date() #fecha de publicación, se convierte a objeto date
    programa = publicacion["programa"] #Nombre del programa de la publicación
    #Se busca que publicacion corresponde al ENSU, si se encuentra, se guarda la fecha y el nombre del programa
    if "Encuesta Nacional de Seguridad Pública Urbana (ENSU)" in programa:
        ensu_fecha = fecha
        ensu_programa = programa

if date.today().isoformat() != ensu_fecha.isoformat():
    print("No es fecha de actualización de datos.")
    exit(0)

print("Hoy es fecha de actualización de datos, se procederá a extraer los datos.")

año = datetime.now().year #año actual
mes = datetime.now().month #mes actual

#Conocemos el trimestre actual
if mes in [1, 2, 3]:
    trimestre = 4
    NomMes = "Diciembre"
    mes = 12
elif mes in [4, 5, 6]:
    trimestre = 1
    NomMes = "Marzo"
    mes = 1
elif mes in [7, 8, 9]:
    trimestre = 2
    NomMes = "Junio"
    mes = 6
elif mes in [10, 11, 12]:
    trimestre = 3
    NomMes = "Septiembre"
    mes = 9

if trimestre == 4:
    año = año - 1 #Si es el trimestre 4, restar un año al año actual

url1 = "https://www.inegi.org.mx/contenidos/programas/ensu/datosabiertos/conjunto_de_datos_ensu_"
url2 = f"{str(año)}_{trimestre}t_csv.zip" #pegar año y trim actual
url = url1 + url2

respuesta = requests.get(url)

#Descargar y extraer el archivo zip
try:
    with zipfile.ZipFile(io.BytesIO(respuesta.content)) as archivo_zip:
        ruta1 = "conjunto_de_datos_ensu_cb_"
        if mes == 12:
            ruta2 = f"{str(mes)}{str(año)[-2:]}/conjunto_de_datos/conjunto_de_datos_ensu_cb_"
            ruta3 = f"{str(mes)}{str(año)[-2:]}.csv"
        else:
            ruta2 = f"0{str(mes)}{str(año)[-2:]}/conjunto_de_datos/conjunto_de_datos_ensu_cb_"
            ruta3 = f"0{str(mes)}{str(año)[-2:]}.csv"
        ruta = ruta1 + ruta2 + ruta3
        archivo_zip.extract(ruta, path=f"datos_ensu_{str(año)}_{trimestre}t")

except zipfile.BadZipFile:
    print("Error: El archivo descargado no es un archivo zip válido.")
    exit(1)

#Leer el archivo CSV extraído
try:
    if mes == 12:
        leer = f"datos_ensu_{str(año)}_{trimestre}t/conjunto_de_datos_ensu_cb_{str(mes)}{str(año)[-2:]}/conjunto_de_datos/conjunto_de_datos_ensu_cb_{str(mes)}{str(año)[-2:]}.csv"
    else:
        leer = f"datos_ensu_{str(año)}_{trimestre}t/conjunto_de_datos_ensu_cb_0{str(mes)}{str(año)[-2:]}/conjunto_de_datos/conjunto_de_datos_ensu_cb_0{str(mes)}{str(año)[-2:]}.csv"
    df = pd.read_csv(leer)
except Exception as e:
    print(f"Error al leer el archivo CSV: {e}")
    exit(1)

#Extraer solo los datos de interés
try:
    df = df[['CVEGEO', 'CVE_ENT', 'NOM_ENT', 'CVE_MUN', 'NOM_MUN', 'BP3_1_06', 'FAC_SEL']]
except KeyError as e:
    print("No se encontró la columna CVEGEO, se creará a partir de CVE_ENT y CVE_MUN.")
    df['CVEGEO'] = df['CVE_ENT'].astype(str).str.zfill(2) + df['CVE_MUN'].astype(str).str.zfill(3)

#Obtener los municipios únicos y sumar si conjunto
columnas = ['CVEGEO', 'NOM_ENT', 'NOM_MUN'] #Columnas usadas en excel
unicos = df.pivot_table(
    index=columnas,
    columns='BP3_1_06',
    values='FAC_SEL',
    aggfunc='sum',
    fill_value=0
).reset_index()

total = (unicos[1]/(unicos[0]+unicos[1])) #Calcular el total de la columna 1
unicos["Total"] = total #Agregar la columna "Total" al DataFrame unicos
unicos.insert(0, "Mes", NomMes) #Agregar columna con el nombre del mes
unicos.insert(1, "Año", año) #Agregar columna con el año
unicos.insert(5, "Conjunto", unicos["NOM_ENT"] + ", " + unicos["NOM_MUN"]) #Agregar columna con el conjunto de Estado y Municipio
unicos.columns = unicos.columns.astype(str) #Convertir los nombres de las columnas a string

Compilado = pd.read_excel("Compilado.xlsx", sheet_name="Compilado") #Leer el archivo Excel compilado
Compilado = Compilado.dropna() #Eliminar filas con valores nulos
final = pd.concat([Compilado, unicos]).reset_index(drop=True) #Concatenar el archivo compilado con los datos actuales

# Escribir el DataFrame
with pd.ExcelWriter("Compilado.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    final.to_excel(writer, sheet_name="Compilado", index=False)

# Crear la tabla
wb = load_workbook("Compilado.xlsx") #Cargar el archivo Excel compilado
ws = wb["Compilado"] #Seleccionar la hoja "Compilado"

ultima_col = get_column_letter(len(final.columns)) #Obtener la letra de la última columna
ultima_fila = len(final) + 1 #Obtener el número de la última fila (sumar 1 para incluir la fila de encabezado)

tabla = Table(
    displayName="ENSU",
    ref=f"A1:{ultima_col}{ultima_fila}"
) #Crear el formato tabla con el nombre "ENSU" y el rango de celdas desde A1 hasta la última celda

estilo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
) #Darle estilo a la tabla

tabla.tableStyleInfo = estilo #Asignar el estilo a la tabla
ws.add_table(tabla) #Agregar la tabla a la hoja "Compilado"

wb.save("Compilado.xlsx") #Guardar el archivo Excel compilado con la tabla creada